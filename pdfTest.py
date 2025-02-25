from google import genai
from markitdown import MarkItDown  # type: ignore
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Border
import re
import argparse
import os
import json
import fitz

#init parser and model
md = MarkItDown()
model = 'gemini-2.0-flash'

#init border style
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def analyze_invoice(invoice_text, client):
    prompt = """
    Ты - опытный бухгалтер, специализирующийся на анализе счетов на оказание услуг.
    Твоя задача - извлечь из предоставленного текста счета ключевую информацию и
    представить её в структурированном JSON формате.

    **Инструкции:**

    1.  **Извлечение информации:** Проанализируй текст и выдели:
        *   **Общая информация о документе:**
            *   "Бухгалтерская справка №" и номер (если есть; иначе "Б.Н.").
            *   Дата "от" (если есть; формат: ДД.ММ.ГГГГ).
            *   Номер счета(документа). формат: "счет № ..."
            *   "по договору №" + номер и дата (если есть; формат: "по договору № ... от ДД.ММ.ГГГГ").
        *   **Исполнитель:**
            *   **Компания-исполнитель**: Полное наименование.
            *   **Адрес исполнителя**:  Адрес (город, улица, дом/офис; **без индекса и области**, формат: г.{Город}, ул.{Улица}, {Дом/офис}. Названия городов и улиц с большой буквы, НЕ капсом).
            *   **УНП исполнителя**: УНП.
            *   **Расчетный счет исполнителя**: (BYxxxxxxxxxxxxxxxxxxxxxxxxxxxx).
            *   **Банк исполнителя**: Название банка и код (через пробел, название банка в ковычках).
        *   **Заказчик:**
            *   **Компания-заказчик**: Полное наименование.
            *   **Адрес заказчика**: (город, улица, дом/офис; **без индекса и области**, формат: г.{Город}, ул.{Улица}, {Дом/офис}).
            *   **УНП заказчика**: УНП.
        *   **Период оказания услуг:** (формат: с ДД.ММ.ГГГГ по ДД.ММ.ГГГГ). Числовой формат.
        *   **Детализация услуг:** (Этот блок должен быть *списком* (list) услуг)
            *   **Наименование услуги(существительное)**: (без "услуги" и "по").
            *   **Сумма без НДС**: (число с разделителем запятой).
            *   **Ставка НДС**: ("-" или "0"/"0,00"  -> "Без НДС").
            *   **Сумма НДС**: (число; при "Без НДС" -> "-").
            *   **Сумма с НДС**: (число).
        *   **Общая стоимость услуг прописью:** (...).
        *   **НДС статус**: ("Без НДС" или "С НДС").
        *   **Кто являеться представителем ЗАКАЗЧИКА**:
            *   **Компания заказчика**: (...).
            *   **Должность представителя**: (...).
            *   **И.О.Фамилия представителя**: (Сначала только инициалы, затем фамилия полностью, без пробелов.).

    2.  **Форматирование JSON, Убедись что весь текст в верном формате:**

        ```json
        {
          "document_info": {
            "document_name": "Бухгалтерская справка № Б.Н.",
            "document_date": "31.12.2024",
            "document_number": "счет № 284"
            "contract_info": "по договору № ПО-103 от 01.09.2019"
          },
          "executor": {
            "company_name": "ООО 'СофтСервис'",
            "address": "г. Гродно, ул. Ленина, 5/2",
            "unp": "500484719",
            "bank_account": "BY22 BELB 3012 1401 7701 1022 6000",
            "bank_name": "ОАО 'БАНК БЕЛВЭБ' BELBBY2X"
          },
          "client": {
            "company_name": "ООО 'ДЕВКРАФТ'",
            "address": "г. Гродно, ул Мостовая, 31",
            "unp": "591007097"
          },
          "service_period": "с 01.12.2024 по 31.12.2024",
          "service_details": [
            {
              "service_name": "Сопровождение программного обеспечения",
              "amount_without_vat": "290,00",
              "vat_rate": "Без НДС",
              "vat_amount": "-",
              "amount_with_vat": "290,00"
            }
          ],
          "total_amount_words": "Двести девяносто белорусских рублей 00 копеек",
          "vat_status": "Без НДС",
          "director": {
            "company_name": "ООО 'ДЕВКРАФТ'",
            "position": "Директор",
            "full_name": "А.В.Яговдик "
          }
        }
        ```
    Текст счета:
    """ + invoice_text

    #analyze document and return result as json
    try:
        response = client.models.generate_content(model=model, contents=[prompt])
        json_text = response.text
        try:
             return json.loads(json_text)
        except json.JSONDecodeError:
            json_text = re.sub(r'```json\n?|```', '', json_text).strip()
            try: return json.loads(json_text)
            except: return json.loads(json_text.replace("'", '"'))
    except Exception as e:
        return f"Error: {e}"

def extract_data_from_analysis(analysis_result):
    data = {}
    if isinstance(analysis_result, str):
        print("Error in analysis:", analysis_result); return data
    
    #extract data from analysis result
    get_val = lambda key, subkey=None: analysis_result.get(key, {}).get(subkey, "") if subkey else analysis_result.get(key, "")
    data["document_name"] = get_val("document_info", "document_name")
    data["document_date"] = get_val("document_info", "document_date")
    data["document_number"] = get_val("document_info", "document_number")
    data["contract_info"] = get_val("document_info", "contract_info")
    data["Исполнитель_Компания"] = get_val("executor", "company_name")
    data["Исполнитель_Адрес"] = get_val("executor", "address")
    data["Исполнитель_УНП"] = get_val("executor", "unp")
    data["Исполнитель_Расчетный_счет"] = get_val("executor", "bank_account")
    bank_name = get_val("executor", "bank_name")
    if bank_name:
      data["Исполнитель_Банк"] = re.sub(r'["“]?БАНК["”]?\b', '"Банк', bank_name, flags=re.IGNORECASE)
    else:
      data["Исполнитель_Банк"] = get_val("executor", "bank_name")
    data["Заказчик_Компания"] = get_val("client", "company_name")
    data["Заказчик_Адрес"] = get_val("client", "address")
    data["Заказчик_УНП"] = get_val("client", "unp")
    data["За период"] = get_val("service_period")
    data["Общая стоимость услуг"] = get_val("total_amount_words")
    data["НДС_Статус"] = get_val("vat_status")
    data["Директор_Должность"] = get_val("director", "position")
    data["Директор_ФИО"] = get_val("director", "full_name")

    data["services"] = analysis_result.get("service_details", [])
    return data

def write_data_to_excel(data, excel_file="output.xlsx"):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        #if file exists, delete and create new (for overwriting)
        wb.close()
        os.remove(excel_file)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

    #write data
    ws.cell(row=2, column=2).value = "Бухгалтерская справка № Б.Н."
    ws.cell(row=2, column=4).value = "от " + data.get("document_date", "")
    ws.cell(row=3, column=2).value = data.get("contract_info", "")
    ws.cell(row=5, column=1).value = "Исполнитель"
    ws.cell(row=5, column=2).value = data.get("Исполнитель_Компания", "")
    executor_details = f'{data.get("Исполнитель_УНП", "")}, {data.get("Исполнитель_Адрес", "")}'
    ws.cell(row=6, column=1).value = executor_details
    ws.cell(row=7, column=1).value = 'Расчетный счет ' + data.get("Исполнитель_Расчетный_счет", "")
    ws.cell(row=8, column=1).value = "в " + data.get("Исполнитель_Банк", "")
    ws.cell(row=10, column=1).value = "Заказчик"
    client_details = f'{data.get("Заказчик_Компания", "")}, {data.get("Заказчик_Адрес", "")}'
    ws.cell(row=10, column=2).value = client_details
    ws.cell(row=12, column=3).value = "За период " + data.get("За период", "") + " оказаны услуги"

    #Table Data
    table_start_row = 15
    ws.merge_cells(start_row=table_start_row - 1, start_column=2, end_row=table_start_row - 1, end_column=5)
    ws.cell(row=table_start_row - 1, column=2).value = "Сумма оказанных услуг"
    ws.cell(row=table_start_row - 1, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=table_start_row - 1, column=2).border = thin_border
    ws.merge_cells(start_row=table_start_row - 1, start_column=1, end_row=table_start_row, end_column=1)
    ws.cell(row=table_start_row - 1, column=1).value = "Наименование услуги"
    ws.cell(row=table_start_row - 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=table_start_row - 1, column=1).border = thin_border

    #write table headers
    table_header = ["Сумма без НДС, руб.", "Ставка НДС", "Сумма НДС, руб.", "Сумма с НДС, руб."]
    for col_num, header in enumerate(table_header, start=2):
        cell = ws.cell(row=table_start_row, column=col_num)
        cell.value = header
        cell.border = thin_border; cell.alignment = Alignment(horizontal='center')

    row_num = table_start_row + 1
    total_without_vat = 0
    total_vat_amount = 0
    total_with_vat = 0

    #write table data
    for service in data.get("services", []):
        service_name_cell = ws.cell(row=row_num, column=1)
        service_name_cell.value = service.get("service_name", "")
        service_name_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row_num, column=2).value = service.get("amount_without_vat", "")
        ws.cell(row=row_num, column=3).value = service.get("vat_rate", "")
        ws.cell(row=row_num, column=4).value = service.get("vat_amount", "")
        ws.cell(row=row_num, column=5).value = service.get("amount_with_vat", "")

        #calculate totals
        try:
            total_without_vat += float(service.get("amount_without_vat", "0").replace(",", "."))
        except ValueError: pass
        try:
            vat_amount_str = service.get("vat_amount", "0").replace(",", ".")
            if vat_amount_str != "-":
                total_vat_amount += float(vat_amount_str)
        except ValueError: pass
        try:
            total_with_vat += float(service.get("amount_with_vat", "0").replace(",", "."))
        except ValueError: pass
        row_num += 1

    ws.cell(row=row_num, column=1).value = "Итого"
    ws.cell(row=row_num, column=2).value = str(total_without_vat).replace(".", ",")
    ws.cell(row=row_num, column=3).value = data.get("НДС_Статус", "Без НДС") if total_without_vat > 0 else ""
    ws.cell(row=row_num, column=4).value = str(total_vat_amount).replace(".", ",") if total_vat_amount > 0 else "-"
    ws.cell(row=row_num, column=5).value = str(total_with_vat).replace(".", ",")

    for row in ws[table_start_row - 1:row_num]:
        for cell in row:
            cell.border = thin_border
            if cell.row > table_start_row and cell.column > 1:
                cell.alignment = Alignment(horizontal='right')

    #Shifting data after the table
    data_start_row = row_num + 2

    ws.cell(row=data_start_row, column=1).value = "Стоимость оказанных услуг: " + data.get("Общая стоимость услуг", "")
    ws.cell(row=data_start_row + 1, column=1).value = data.get("НДС_Статус", "")
    ws.cell(row=data_start_row + 3, column=1).value = "Документ составлен в единоличном порядке в соответствии с п.1.7. Договора, на основании п. 1 "
    ws.cell(row=data_start_row + 4, column=1).value = "Постановления Министерства Финансов Республики Беларусь от 12.02.2018 № 13"
    ws.cell(row=data_start_row + 6, column=1).value = "Основание: " + data.get("document_number", "") + " от " + data.get("document_date", "")
    ws.cell(row=data_start_row + 9, column=1).value = f"{data.get('Директор_Должность', '')} {data.get('Заказчик_Компания', '')}"
    ws.cell(row=data_start_row + 9, column=5).value = data.get("Директор_ФИО", "")
    ws.cell(row=data_start_row + 9, column=3).border = Border(bottom=Side(style='thin'))

    #set column widths
    column_widths = {'A': 22, 'B': 15, 'C': 15, 'D': 15, 'E': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width + 1

    wb.save(excel_file)
    print(f"Data written to {excel_file}")

def main():
    parser = argparse.ArgumentParser(description="Analyze an invoice and extract data to an Excel file.")

    parser.add_argument("-i", "--input", required=True, help="Path to the input PDF file.")
    parser.add_argument("-k", "--key", required=True, help="Your Google Gemini API key.")

    args = parser.parse_args()

    client = genai.Client(api_key=args.key)

    if not os.path.exists(args.input):
        print(f"Error: Input file '{args.input}' not found.")
        return

    # Открываем PDF
    doc = fitz.open(args.input)
    total_pages = len(doc)

    print(f"Оригинальный PDF содержит {total_pages} страниц.")

    while True:
        try:
            num_pages = int(input(f"Сколько страниц оставить? (0 - оставить все, 1-{total_pages}): "))
            if 0 <= num_pages <= total_pages:
                break
            else:
                print("Ошибка: число вне диапазона.")
        except ValueError:
            print("Ошибка: введите целое число.")

    if num_pages == 0:
        print("Используется оригинальный PDF без изменений.")
        pdf_to_process = args.input
    else:
        # Создаём временный PDF
        temp_pdf = "temp.pdf"
        new_doc = fitz.open()

        for page_num in range(num_pages):
            new_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)

        new_doc.save(temp_pdf)
        new_doc.close()
        pdf_to_process = temp_pdf
        print(f"Создан временный PDF: {temp_pdf}")

    doc.close()

    # Обрабатываем PDF (оригинальный или временный)
    result = md.convert(pdf_to_process)

    analysis_result = analyze_invoice(result.text_content, client)
    print("\nРезультат анализа:")
    print(analysis_result)

    extracted_data = extract_data_from_analysis(analysis_result)

    base_name = os.path.splitext(args.input)[0]
    output_file = f"{base_name}.xlsx"

    write_data_to_excel(extracted_data, output_file)

    if num_pages != 0:
        os.remove(temp_pdf)
        print(f"Временный PDF {temp_pdf} удалён.")


if __name__ == "__main__":
    main()